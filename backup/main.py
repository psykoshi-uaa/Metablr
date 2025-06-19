from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
import sys
import csv
import string

class Header:
	def __init__(self, header_list):
		self.header = header_list
		self.name_ind = 0
		self.areamax_ind = 0
		self.normarea_ind_start = 0
		self.normarea_ind_end = 0
		self.rsd_ind = 0

		self.autoset_name_ind()
		self.autoset_areamax_ind()
		self.autoset_normarea_ind()
		self.autoset_rsd_ind()


	def autoset_areamax_ind(self):
		counter = 0
		for column in self.header:
			if column == "Area (Max.)":
				self.areamax_ind = counter
				return
			counter += 1


	def autoset_name_ind(self):
		counter = 0
		for column in self.header:
			if column == "Name":
				self.name_ind = counter
				return
			counter += 1


	def autoset_rsd_ind(self):
		counter = 0
		for column in self.header:
			if column == "RSD Corr. QC Areas [%]":
				self.rsd_ind = counter
				return
			counter += 1


	def autoset_normarea_ind(self):
		is_start = False
		counter = 0
		normarea_counter = 0
		for column in self.header:
			if "Norm. Area:" in column:
				if "QC" in column:
					if self.normarea_ind_start == 0:
						self.normarea_ind_start = counter
					normarea_counter += 1
			counter += 1
		
		self.normarea_ind_end = self.normarea_ind_start + normarea_counter - 1


	def get_name_ind(self):
		return self.name_ind


	def get_areamax_ind(self):
		return self.areamax_ind

	
	def get_normarea_ind_start(self):
		return int(self.normarea_ind_start)


	def get_normarea_ind_end(self):
		return int(self.normarea_ind_end)


	def get_rsd_ind(self):
		return self.rsd_ind



class Metabolite:
	def __init__(self, data_list, name_ind, areamax_ind, normarea_ind_start, normarea_ind_end, rsd_ind):
		self.data = data_list
		self.name = data_list[name_ind]
		self.areamax = data_list[areamax_ind]
		self.rsd = data_list[rsd_ind]
		self.avg_normarea = 0
		self.autoset_avg_normarea(normarea_ind_start, normarea_ind_end)
	

	def __eq__(self, other):
		if self.name == other.get_name():
			return True
		return False	


	def __gt__(self, other):
		if self.areamax > other.get_areamax():
			return True
		return False


	def __lt__(self, other):
		if self.get_areamax() < other.get_areamax():
			return True
		return False


	def __gte__(self, other):
		if self.get_areamax() >= other.get_areamax():
			return True
		return False


	def __lte__(self, other):
		if self.get_areamax() >= other.get_areamax():
			return True
		return False


	def get_name(self):
		return self.name


	def get_areamax(self):
		return float(self.areamax)


	def get_areamax_as_str(self):
		return self.areamax


	def get_rsd(self):
		return float(self.rsd)


	def get_rsd_as_str(self):
		return self.rsd

	
	def get_avg_normarea(self):
		return float(self.avg_normarea)


	def get_avg_normarea_as_str(self):
		return str(self.avg_normarea)


	def autoset_avg_normarea(self, start_ind, end_ind):
		max_normarea = float(0)
	
		for i in range(start_ind, end_ind + 1):
			max_normarea += float(self.data[i])
			print(self.data[i])

		self.avg_normarea = max_normarea / (end_ind - start_ind)
		print()



class Metabolomics:
	def __init__(self, csv_filename):
		self.all_metabolites = []
		with open(csv_filename, 'r') as csv_file:
			csv_reader = csv.reader(csv_file)
			counter = 0

			for data in csv_reader:
				if counter == 0:
					self.header = Header(data)
				else:
					self.all_metabolites.append(Metabolite(data, self.header.get_name_ind(), self.header.get_areamax_ind(), self.header.get_normarea_ind_start(), self.header.get_normarea_ind_end(), self.header.get_rsd_ind()))
				counter += 1
		self.list_size = counter - 1
		csv_file.close()


	def __gt__(self, other):
		if self.list_size > other.get_list_size():
			return True
		return False	


	def get_header(self):
		return self.header.copy()


	def get_all_metabolites(self):
		return self.all_metabolites.copy()
	

	def get_names(self):
		for metabolite in self.all_metabolites:
			print(metabolite.get_name())		#DEBUG PRINT


	def get_areamaxi(self):
		for metabolite in self.all_metabolites:
			print(metabolite.get_areamax())		#DEBUG PRINT


	def get_list_size(self):
		return self.list_size


	def print_name_normarea_rsd(self):
		counter = 0
		for metabolite in self.get_all_metabolites():
			if counter % 2 == 0:
				print("\033[39m", metabolite.get_name().ljust(35, ' '), metabolite.get_avg_normarea_as_str().ljust(20, ' '),  metabolite.get_rsd_as_str().ljust(5, ' '), "\033[0m")
			else:
				print("\033[39;104m", metabolite.get_name().ljust(35, ' '), metabolite.get_avg_normarea_as_str().ljust(20, ' '),  metabolite.get_rsd_as_str().ljust(5, ' '), "\033[0m")
			counter += 1
		

	def stitch(self, other):
		temp_metabolites = self.get_all_metabolites()
		counter = 0

		for metabolite in temp_metabolites:
			for metabolite_other in other.get_all_metabolites():
				if metabolite_other == metabolite:
					if metabolite < metabolite_other:
						self.all_metabolites[counter] = metabolite_other
			counter += 1
						
		for metabolite_other in other.get_all_metabolites():
			is_unique = True
			for metabolite in self.get_all_metabolites():
				if metabolite_other == metabolite:
					is_unique = False

			if is_unique == True:
				self.all_metabolites.append(metabolite_other)

		self.sort_list()


	def sort_list(self):
		sorted_names = []
		temp_metabolites = []

		for metabolite in self.get_all_metabolites():
			sorted_names.append(metabolite.get_name())
		sorted_names.sort()

		for name in sorted_names:
			for metabolite in self.get_all_metabolites():
				if metabolite.get_name() == name:
					temp_metabolites.append(metabolite)
					break

		self.all_metabolites[:] = []
		self.all_metabolites = temp_metabolites


#def create_metabolomics_xls(metabolomics):


def create_name_exclusive_csv(pos, neg, both, header):
	def_cell_color = PatternFill(start_color="e26b0a", end_color="e26b0a", fill_type = "solid")	
	pos_cell_color = PatternFill(start_color="538dd5", end_color="538dd5", fill_type = "solid")	
	neg_cell_color = PatternFill(start_color="92d050", end_color="92d050", fill_type = "solid")	
	bold_font = Font(bold=True)
	data = []
	header = ["Positive Mode", "Negative Mode", "Both Modes", "Combined for MetaboAnalyst"]
	pos_counter = 0
	neg_counter = 0

	metabo_color_list = []

	for metabolite in both.get_all_metabolites():
		cell_color = def_cell_color
		is_pos = False
		is_neg = False
		line = ['','','','']

		pos_name = pos.get_all_metabolites()[pos_counter].get_name()
		neg_name = neg.get_all_metabolites()[neg_counter].get_name()
		pos_areamax = pos.get_all_metabolites()[pos_counter].get_areamax()
		neg_areamax = neg.get_all_metabolites()[neg_counter].get_areamax()

		if pos_name == metabolite.get_name():
			if pos_areamax == metabolite.get_areamax():
				cell_color = pos_cell_color
			line[0] = metabolite.get_name()
			is_pos = True
			pos_counter += 1

		if neg_name == metabolite.get_name():
			if neg_areamax == metabolite.get_areamax() and not pos_areamax == metabolite.get_areamax():
				cell_color = neg_cell_color
			line[1] = metabolite.get_name()
			is_neg = True
			neg_counter += 1

		if is_pos and is_neg:
			line[2] = metabolite.get_name()
		line[3] = metabolite.get_name()
		data.append(line)
		metabo_color_list.append(cell_color)
	
	wb = Workbook()
	ws = wb.active
	ws.title = "metabolite_list"
	ws.append(header)

	for col in ws["A:D"]:
		for cell in col:
			cell.font = bold_font

	#for row in range(ws.min_row, ws.max_row + 1):

	for line in data:
		ws.append(line)

	counter = 0
	outline = Side(style = "thin", color = "000000")
	for row in ws.iter_rows(min_row=2, min_col = 4, max_col=4, max_row=ws.max_row):
		for cell in row:
			cell.fill = metabo_color_list[counter]
			cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
		counter += 1


	dim_holder = DimensionHolder(worksheet=ws)
	for col in range(ws.min_column, ws.max_column + 1):
		dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=30)
	dim_holder['F'] = ColumnDimension(ws, min=ws.max_column + 2, max=ws.max_column + 2, width=15)
	ws.column_dimensions = dim_holder

	cell = ws["F2"]
	cell.value = "positive"
	cell.fill = pos_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "right", vertical = "bottom")

	cell = ws["F3"]
	cell.value = "negative"
	cell.fill = neg_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "right", vertical = "bottom")

	cell = ws["F4"]
	cell.value = "determined"
	cell.fill = def_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "right", vertical = "bottom")

	wb.save("metabolite_list.xlsx")



if __name__ == "__main__":
	pos_file = "formatted.csv"
	#neg_file = "negative.csv"

	metabolomics_pos = Metabolomics(pos_file)
	#metabolomics_neg = Metabolomics(neg_file)
	#metabolomics_both = Metabolomics(pos_file)

	#metabolomics_both.stitch(metabolomics_neg)
	metabolomics_pos.print_name_normarea_rsd()
	#create_name_exclusive_csv(metabolomics_pos, metabolomics_neg, metabolomics_both)
