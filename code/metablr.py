import sys
import os.path
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import colors, Font, Color, PatternFill, Border, Side, Alignment


HELP_SEG = "\nAppend a positive and then a negative xlsx file respectively followed by the name of the file you are creating\nlist of Metablr commands:\n\t-S | --summary\tsummarized two appended metabolomics files into a combined summary table\n\t-R | --reformat\treformats the appended files into a combined data table\n\t-h | --help\thi\n"
ERROR_0 = "Too few xlsx files"
ERROR_1 = "Too many flags"
ERROR_2 = "Too many arguments"
CLI_SAVE_AS = "PL4C3h0Ld3r51l3n4mE"
WORKBOOK_SHEET_NAME = "Compounds"


class Program_Log():
	def __init__(self):
		self.reg_log = []
		self.error_log = []
		self.error_log.append("\n\033[31;1m\t----ERROR----\033[0m\n")
	
	
	def get_error_count(self):
		return len(self.error_log) - 1

	def append_reg(self, reg):
		self.reg_log.append(reg)

	def append_error(self, error):
		self.error_log.append(error)

	def print_log(self):
		for line in self.reg_log:
			print(line)
		if (len(self.error_log) > 1):
			self.error_log.append("\n\033[31;1m\t--END ERROR--\033[0m\n")
			for line in self.error_log:
				print(line)



class Headers():
	def __init__(self, xlsx_filename, program_log):
		self.headers = { "name":"name", "rsd":"rsd corr. qc areas [%]", "norm":"norm. area:", "repl":"replicate grouped area:" }
		self.ind_name = 0
		self.ind_normarea_start = 0
		self.ind_normarea_end = 0
		self.ind_repl_start = 0
		self.ind_repl_end = 0
		self.ind_rsd = 0
		self.autoset_all_ind(xlsx_filename, program_log)

	
	def autoset_all_ind(self, xlsx_filename, program_log):
		norm_ind_counter = 0
		repl_ind_counter = 0
		wb = openpyxl.load_workbook(xlsx_filename)
		ws = wb.worksheets[0]
		for col in range(1, ws.max_column + 1):
			cell = ws.cell(row = 1, column = col)
			if (self.get_header_title("name", program_log) in cell.value.lower()):
				self.set_ind_name(col - 1)
			if (self.get_header_title("rsd", program_log) in cell.value.lower()):
				self.set_ind_rsd(col - 1)
			if (self.get_header_title("norm", program_log) in cell.value.lower()):
				if (norm_ind_counter == 0):
					self.set_ind_normarea_start(col - 1)
				if ("qc" in cell.value.lower()):				#FIX MAGIC STRING
					norm_ind_counter += 1
			if (self.get_header_title("repl", program_log) in cell.value.lower()):
				if (repl_ind_counter == 0):
					self.set_ind_repl_start(col)
				repl_ind_counter += 1

		if (norm_ind_counter > 0):
			self.set_ind_normarea_end(self.get_ind_normarea_start() + norm_ind_counter - 1)
		if (repl_ind_counter > 0):
			self.set_ind_repl_end(self.get_ind_repl_start() + repl_ind_counter - 1)


	def get_ind_name(self):
		return self.ind_name


	def get_ind_normarea_start(self):
		return self.ind_normarea_start


	def get_ind_normarea_end(self):
		return self.ind_normarea_end


	def get_ind_repl_start(self):
		return self.ind_repl_start


	def get_ind_repl_end(self):
		return self.ind_repl_end


	def get_ind_rsd(self):
		return self.ind_rsd


	def get_header_title(self, title, program_log):
		try:
			return self.headers[title]
		except Exception as e:
			program_log.append_error(e)
			return ""


	def set_ind_name(self, ind):
		self.ind_name = ind
				
		
	def set_ind_normarea_start(self, ind):
		self.ind_normarea_start = ind
				
		
	def set_ind_normarea_end(self, ind):
		self.ind_normarea_end = ind
				
		
	def set_ind_repl_start(self, ind):
		self.ind_repl_start = ind


	def set_ind_repl_end(self, ind):
		self.ind_repl_end = ind


	def set_ind_rsd(self, ind):
		self.ind_rsd = ind
				
		

class Metabolite():
	def __init__(self, data, name_ind, rsd_ind, normarea_start_ind, normarea_end_ind):
		self.data = data
		self.name = self.get_data()[name_ind]
		self.rsd = self.get_data()[rsd_ind]
		self.avg_normarea = 0
		self.calculate_avg_normarea(normarea_start_ind, normarea_end_ind)


	def __eq__(self, other):
		if (self.get_name() == other.get_name()):
			return True
		return False


	def __gt__(self, other):
		if (self.get_rsd() > other.get_rsd()):
		#if (self.get_avg_normarea()/self.get_rsd() > other.get_avg_normarea()/self.get_rsd()):
			return True
		return False


	def get_data(self):
		return self.data

		
	def get_name(self):
		return self.name


	def get_rsd(self):
		return self.rsd


	def get_avg_normarea(self):
		return self.avg_normarea


	def get_data_at_ind(self, ind):
		return self.data[ind]
	

	def set_name(self, name):
		self.name = name


	def set_rsd(self, rsd):
		self.rsd = rsd


	def calculate_avg_normarea(self, start_ind, end_ind):
		temp_normarea_total = 0
		avg_normarea = 0

		for i in range(start_ind, end_ind + 1):
			temp_normarea_total += self.data[i]

		avg_normarea = temp_normarea_total / (end_ind - start_ind + 1)
		self.avg_normarea = avg_normarea



class Metabolomics():
	def __init__(self, xlsx_filename, program_log):
		self.filename = xlsx_filename
		self.headers = Headers(xlsx_filename, program_log)
		self.metabolites = []
		self.repl_group_names = []
		self.repl_sample_names = []
		self.row_size = 0
		self.col_size = 0
		self.num_metabolites = 0
		self.autoset(xlsx_filename)


	def get_filename(self):
		return self.filename


	def get_headers(self):
		return self.headers

	
	def get_row_size(self):
		return self.row_size

	
	def get_col_size(self):
		return self.col_size

	
	def get_metabolites(self):
		return self.metabolites


	def get_names(self):
		names = []
		for metabolite in self.metabolites:
			names.append(metabolite.get_name())
		return names


	def get_data_at_ind(self, ind):
		data = []
		for metabolite in self.metabolites:
			data.append(metabolite.get_data_at_ind(ind))
		return data


	def get_sample_name(self, ind, program_log):
		try:
			return self.repl_sample_names[ind]
		except Exception as e:
			program_log.append_error(e)
			return ""
			
	def get_group_name(self, ind, program_log):
		try:
			return self.repl_group_names[ind]
		except Exception as e:
			program_log.append_error(e)
			return ""
			
	def get_num_metabolites(self):
		return self.num_metabolites


	def autoset(self, xlsx_filename):
		wb = openpyxl.load_workbook(xlsx_filename)
		ws = wb[WORKBOOK_SHEET_NAME]
		self.num_metabolites = ws.max_row - 1

		row_size = ws.max_row
		col_size = ws.max_column
		
		for row in range(2, ws.max_row + 1):
			metabolite_data = []
			for col in range(1, ws.max_column + 1):
				cell = ws.cell(row = row, column = col)
				metabolite_data.append(cell.value)

			self.metabolites.append(Metabolite(metabolite_data, self.headers.get_ind_name(), self.headers.get_ind_rsd(), self.headers.get_ind_normarea_start(), self.headers.get_ind_normarea_end()))


	def autoset_repl_names(self, cat_vars):
		sample_name = "PLACEHOLDER"
		temp_name = ""
		sample_num = 1
		for var in cat_vars:
			sample_name = var
			if (temp_name == var):
				sample_num += 1
			else:
				sample_num = 1
				temp_name = var

			sample_name += (" " + (str(sample_num).zfill(3)))
			self.repl_group_names.append(var)
			self.repl_sample_names.append(sample_name)
						

	def stitch_with(self, other):
		counter = 0
		for other_metab in other.get_metabolites().copy():
			is_unique = True
			for self_metab in self.get_metabolites().copy():
				if (other_metab == self_metab):
					is_unique = False
					if (other_metab.get_rsd() < self_metab.get_rsd()):
						self.get_metabolites()[counter] = other_metab
						break
					elif (other_metab.get_rsd() == self_metab.get_rsd()):
						if (other_metab.get_avg_normarea() > self_metab.get_avg_normarea()):
							self.get_metabolites()[counter] = other_metab
				counter += 1

			if (is_unique == True):
				self.metabolites.append(other_metab)

			counter = 0

		self.sort_list()


	def sort_list(self):
		sorted_names = []
		temp_metabolites = []

		for metabolite in self.get_metabolites().copy():
			sorted_names.append(metabolite.get_name())
		sorted_names.sort()

		for name in sorted_names:
			for metabolite in self.get_metabolites().copy():
				if (metabolite.get_name() == name):
					temp_metabolites.append(metabolite)
					break

		self.metabolites[:] = []
		self.metabolites = temp_metabolites

	
	def print_fancy(self):
		counter = 0
		for metabolite in self.metabolites:
			color = "\033[39;104m"
			if (counter % 2 == 0):
				color = "\033[39m"
			print(color, metabolite.get_name().ljust(35, ' '), str(metabolite.get_avg_normarea()).ljust(20, ' '), str(metabolite.get_rsd()).ljust(20, ' '), str(metabolite.get_avg_normarea()/metabolite.get_rsd()), "\033[0m")
			counter += 1



def get_input_file_cat_vars(input_filename):
	wb = openpyxl.load_workbook(input_filename)
	ws = wb[wb.sheetnames[0]]
	samp_ind = 0
	cat_var_ind = 0
	cat_var_list = []
	for col in range(ws.min_column, ws.max_column):
		cell = ws.cell(row=1, column=col)
		if ("sample type" in cell.value.lower()):
			samp_ind = col
		elif ("categorical" in cell.value.lower() and "variable" in cell.value.lower()):
			cat_var_ind = col

	for row in range(ws.min_row + 1, ws.max_row + 1):
		is_sample = False
		for col in range(samp_ind, cat_var_ind + 1):
			cell = ws.cell(row=row, column=col)
			if (col == samp_ind and "sample" in cell.value.lower()):
				is_sample = True
					
			if (col == cat_var_ind and is_sample == True):
				cat_var_list.append(cell.value)

	return cat_var_list


def create_reformatted_xlsx_file(metab1, metab2, stitched, save_as, program_log):
		#gray
	def_cell_color = PatternFill(start_color="868686", end_color="868686", fill_type = "solid")	
		#orange
	metab1_cell_color = PatternFill(start_color="e97700", end_color="e97700", fill_type = "solid")	
		#purple
	metab2_cell_color = PatternFill(start_color="b971f0", end_color="b971f0", fill_type = "solid")	
	outline = Side(style = "thin", color = "000000")
	wb = Workbook()
	ws = wb.active
	ws.title = WORKBOOK_SHEET_NAME
	header_names = ["Name", "Group"]

	for name in stitched.get_names():
		header_names.append(name)
	ws.append(header_names)

	counter = 3
	metab1_counter = 0
	metab2_counter = 0
	for metabolite in stitched.get_metabolites().copy():
		metab1_name = metab1.get_metabolites()[metab1_counter].get_name()
		metab2_name = metab2.get_metabolites()[metab2_counter].get_name()
		metab1_normarea = metab1.get_metabolites()[metab1_counter].get_avg_normarea()
		metab2_normarea = metab2.get_metabolites()[metab2_counter].get_avg_normarea()

		if metab1_name == metabolite.get_name():
			if metab1_normarea == metabolite.get_avg_normarea():
				cell_color = metab1_cell_color
			metab1_counter += 1

		if metab2_name == metabolite.get_name():
			if metab2_normarea == metabolite.get_avg_normarea() and not metab1_normarea == metabolite.get_avg_normarea():
				cell_color = metab2_cell_color
			metab2_counter += 1

		cell = ws.cell(row=1, column=counter)
		cell.fill = cell_color
		cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
		cell.alignment = Alignment(horizontal = "right", vertical = "bottom")
		counter += 1

	counter = 0
	repl_ind_start = stitched.get_headers().get_ind_repl_start()
	repl_ind_end = stitched.get_headers().get_ind_repl_end() + 1
	for row in range(repl_ind_start - 1, repl_ind_end - 1):
		data = []
		data.append(stitched.get_sample_name(counter, program_log))
		data.append(stitched.get_group_name(counter, program_log))
		for repl_data in stitched.get_data_at_ind(row):
			data.append(repl_data)
		ws.append(data)
		counter += 1

	dim_holder = DimensionHolder(worksheet=ws)
	for col in range(ws.min_column + 2, ws.max_column + 1):
		dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)		#MAGIC NUMBER
	ws.column_dimensions = dim_holder

	wb.save(save_as)
	


def create_stitched_xlsx_file(metab1, metab2, stitched, save_as):
		#gray
	def_cell_color = PatternFill(start_color="868686", end_color="868686", fill_type = "solid")	
		#orange
	metab1_cell_color = PatternFill(start_color="e97700", end_color="e97700", fill_type = "solid")	
	metab1_cell_color_light = PatternFill(start_color="fcd5b4", end_color="fcd5b4", fill_type = "solid")	
		#purple
	metab2_cell_color = PatternFill(start_color="b971f0", end_color="b971f0", fill_type = "solid")	
	metab2_cell_color_light = PatternFill(start_color="ccc0da", end_color="ccc0da", fill_type = "solid")	

	bold_font = Font(bold=True)
	data = []
	header = ["Combined for MetaboAnalyst", "Both", "Positive RSD QC", "Positive QC Norm Avg.", "Negative RSD QC", "Negative QC Norm Avg."]
	metab1_counter = 0
	metab2_counter = 0

	metabo_color_list = []

	for metabolite in stitched.get_metabolites().copy():
		cell_color = def_cell_color
		is_metab1 = False
		is_metab2 = False
		line = ['','','','','','']

		metab1_name = metab1.get_metabolites()[metab1_counter].get_name()
		metab2_name = metab2.get_metabolites()[metab2_counter].get_name()
		metab1_rsd = metab1.get_metabolites()[metab1_counter].get_rsd()
		metab2_rsd = metab2.get_metabolites()[metab2_counter].get_rsd()
		metab1_normarea = metab1.get_metabolites()[metab1_counter].get_avg_normarea()
		metab2_normarea = metab2.get_metabolites()[metab2_counter].get_avg_normarea()

		line[0] = metabolite.get_name()
		if metab1_name == metabolite.get_name():
			if metab1_normarea == metabolite.get_avg_normarea():
				cell_color = metab1_cell_color
			line[2] = metab1_rsd
			line[3] = metab1_normarea
			is_metab1 = True
			metab1_counter += 1

		if metab2_name == metabolite.get_name():
			if metab2_normarea == metabolite.get_avg_normarea() and not metab1_normarea == metabolite.get_avg_normarea():
				cell_color = metab2_cell_color
			line[4] = metab2_rsd
			line[5] = metab2_normarea
			is_metab2 = True
			metab2_counter += 1

		if is_metab1 and is_metab2:
			line[1] = 'X'
		data.append(line)
		metabo_color_list.append(cell_color)
	
	wb = Workbook()
	ws = wb.active
	ws.title = WORKBOOK_SHEET_NAME
	ws.append(header)

	for col in ws["A:G"]:
		for cell in col:
			cell.font = bold_font
			cell.alignment = Alignment(horizontal = "left", vertical = "bottom")
			

	for line in data:
		ws.append(line)

	counter = 0
	outline = Side(style = "thin", color = "000000")
	for row in ws.iter_rows(min_row=2, min_col = 1, max_col=1, max_row=ws.max_row):
		for cell in row:
			cell.fill = metabo_color_list[counter]
			cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
		counter += 1

	for row in ws.iter_rows(min_row=2, min_col = 2, max_col=6, max_row=ws.max_row):
		for cell in row:
			cell.alignment = Alignment(horizontal = "center", vertical = "bottom")
		counter += 1

	for row in ws.iter_rows(min_row=2, min_col = 3, max_col=4, max_row=ws.max_row):
		for cell in row:
			cell.fill = metab1_cell_color_light
			cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
		counter += 1

	for row in ws.iter_rows(min_row=2, min_col = 5, max_col=6, max_row=ws.max_row):
		for cell in row:
			cell.fill = metab2_cell_color_light
			cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
		counter += 1

	dim_holder = DimensionHolder(worksheet=ws)
	for col in range(ws.min_column, ws.max_column + 1):
		dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=10)
	dim_holder['A'] = ColumnDimension(ws, min=1, max=1, width=40)
	dim_holder['D'] = ColumnDimension(ws, min=4, max=4, width=15)
	dim_holder['F'] = ColumnDimension(ws, min=6, max=6, width=15)
	dim_holder['H'] = ColumnDimension(ws, min=8, max=8, width=25)
	ws.column_dimensions = dim_holder

	cell = ws["C1"]
	cell.fill = metab1_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "right", vertical = "bottom")
	cell = ws["D1"]
	cell.fill = metab1_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "right", vertical = "bottom")

	cell = ws["E1"]
	cell.fill = metab2_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "right", vertical = "bottom")
	cell = ws["F1"]
	cell.fill = metab2_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "right", vertical = "bottom")

	cell = ws["H2"]
	cell.value = "Positive Mode"
	cell.fill = metab1_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "center", vertical = "bottom")
	cell = ws["H3"]
	cell.value = "Negative Mode"
	cell.fill = metab2_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "center", vertical = "bottom")

	wb.save(save_as)



def check_filename(xlsx_filename, program_log):
	try:
		wb = openpyxl.load_workbook(xlsx_filename)
		return 1
	except Exception as e:
		program_log.append_error(e)
		return 0



def program_state(args, program_log):
	stitch_mode = False
	reformat_mode = False

	xlsx_files = []
	flags = []
	for argv in args:
		if (argv == "-S") or (argv == "--stitch"):
			stitch_mode = True
			flags.append(argv)
		elif (argv == "-R") or (argv == "--reformat"):
			reformat_mode = True
			flags.append(argv)
		elif (argv == "-h") or (argv == "--help"):
			program_log.append_reg(HELP_SEG)
			flags.append(argv)
			return
		elif (".xlsx" in argv and argv != args[-1]):
			# if (check_filename(argv, program_log) == 1):
			xlsx_files.append(argv)

	if (len(flags) == 0):
		program_log.append_reg("Please type -h to see a list of commands")
		return
	if (len(flags) != 1):
		program_log.append_error(ERROR_1)
		return

	if (stitch_mode == True):
		if (len(xlsx_files) < 2):
			program_log.append_error(ERROR_0)
			return
		elif (len(xlsx_files) > 4):
			program_log.append_error(ERROR_2)
			return
		metab1 = Metabolomics(xlsx_files[0], program_log)
		metab2 = Metabolomics(xlsx_files[1], program_log)
		metab_stitched = Metabolomics(xlsx_files[0], program_log)
		metab_stitched.stitch_with(Metabolomics(xlsx_files[1], program_log))

		create_stitched_xlsx_file(metab1, metab2, metab_stitched, args[-1])

		if (os.path.isfile(args[-1])):
			program_log.append_reg(xlsx_files[0] + " + " + xlsx_files[1]  + "\033[32;1m ---> \033[33;1m\"" + args[-1] + "\"\033[0m...")
		else:
			program_log.append_reg(xlsx_files[0] + " + " + xlsx_files[1]  + "\033[31;1m ---X \033[33;1m\"" + args[-1] + "\"\033[0m...")
		
	if (reformat_mode == True):
		if (len(xlsx_files) < 4):
			program_log.append_error(ERROR_0)
			program_log.append_error("Ensure you are using 4 xlsx files for data table export")
			return
		elif (len(xlsx_files) > 4):
			program_log.append_error(ERROR_2)
			program_log.append_error("Ensure you are using 4 xlsx files for data table export")
			return
		inp_file1 = get_input_file_cat_vars(xlsx_files[1])
		metab1 = Metabolomics(xlsx_files[0], program_log)
		metab1.autoset_repl_names(inp_file1)

		inp_file2 = get_input_file_cat_vars(xlsx_files[3])
		metab2 = Metabolomics(xlsx_files[2], program_log)
		metab2.autoset_repl_names(inp_file2)

		metab_stitched = Metabolomics(xlsx_files[0], program_log)
		metab_stitched.stitch_with(Metabolomics(xlsx_files[2], program_log))
		metab_stitched.autoset_repl_names(inp_file1)

		create_reformatted_xlsx_file(metab1, metab2, metab_stitched, args[-1], program_log)
	
		if (os.path.isfile(args[-1])):
			program_log.append_reg("Pos + Neg data" + "\033[32;1m ---> \033[33;1m\"" + args[-1] + "\"\033[0m...")
		else:
			program_log.append_error("Pos + Neg data" + "\033[31;1m ---X \033[33;1m\"" + args[-1] + "\"\033[0m...")



if __name__ == "__main__":
	program_log = Program_Log()

	program_state(sys.argv, program_log)

	program_log.print_log()
