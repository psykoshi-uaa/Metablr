import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import colors, Font, Color, PatternFill, Border, Side, Alignment


WORKBOOK_SHEET_NAME = "Compounds"


class Headers():
	def __init__(self, xlsx_filename):
		self.headers = { "name":"name", "rsd":"rsd corr. qc areas [%]", "norm":"norm. area:", "repl":"replicate grouped area:" }
		self.ind_name = 0
		self.ind_normarea_start = 0
		self.ind_normarea_end = 0
		self.ind_repl_start = 0
		self.ind_repl_end = 0
		self.ind_rsd = 0
		self.autoset_all_ind(xlsx_filename)

	
	def autoset_all_ind(self, xlsx_filename):
		norm_ind_counter = 0
		repl_ind_counter = 0
		wb = openpyxl.load_workbook(xlsx_filename)
		ws = wb[WORKBOOK_SHEET_NAME]
		for col in range(1, ws.max_column + 1):
			cell = ws.cell(row = 1, column = col)
			if (self.get_header_title("name") in cell.value.lower()):
				self.set_ind_name(col - 1)
			if (self.get_header_title("rsd") in cell.value.lower()):
				self.set_ind_rsd(col - 1)
			if (self.get_header_title("norm") in cell.value.lower()):
				if (norm_ind_counter == 0):
					self.set_ind_normarea_start(col - 1)
				if ("qc" in cell.value.lower()):				#FIX MAGIC STRING
					norm_ind_counter += 1
			if (self.get_header_title("repl") in cell.value.lower()):
				if (repl_ind_counter == 0):
					self.set_ind_repl_start(col)
				repl_ind_counter += 1

		if (norm_ind_counter > 0):
			self.set_ind_normarea_end(self.get_ind_normarea_start() + norm_ind_counter - 1)
		if (repl_ind_counter > 0):
			self.set_ind_repl_end(self.get_ind_repl_start() + repl_ind_counter - 1)


	def get_ind_name(self):
		return self.ind_name


	def get_ind_normarea_start(self):
		return self.ind_normarea_start


	def get_ind_normarea_end(self):
		return self.ind_normarea_end


	def get_ind_repl_start(self):
		return self.ind_repl_start


	def get_ind_repl_end(self):
		return self.ind_repl_end


	def get_ind_rsd(self):
		return self.ind_rsd


	def get_header_title(self, title):
		try:
			return self.headers[title]
		except Exception as e:
			print("error: get_header_file(title); title not found in Header.headers, returning nothing")
			return ""


	def set_ind_name(self, ind):
		self.ind_name = ind
				
		
	def set_ind_normarea_start(self, ind):
		self.ind_normarea_start = ind
				
		
	def set_ind_normarea_end(self, ind):
		self.ind_normarea_end = ind
				
		
	def set_ind_repl_start(self, ind):
		self.ind_repl_start = ind


	def set_ind_repl_end(self, ind):
		self.ind_repl_end = ind


	def set_ind_rsd(self, ind):
		self.ind_rsd = ind
				
		

class Metabolite():
	def __init__(self, data, name_ind, rsd_ind, normarea_start_ind, normarea_end_ind):
		self.data = data
		self.name = self.get_data()[name_ind]
		self.rsd = self.get_data()[rsd_ind]
		self.avg_normarea = 0
		self.calculate_avg_normarea(normarea_start_ind, normarea_end_ind)


	def __eq__(self, other):
		if (self.get_name() == other.get_name()):
			return True
		return False


	def __gt__(self, other):
		if (self.get_avg_normarea() > other.get_avg_normarea()):
			return True
		return False


	def get_data(self):
		return self.data

		
	def get_name(self):
		return self.name


	def get_avg_normarea(self):
		return self.avg_normarea


	def get_data_at_ind(self, ind):
		return self.data[ind]
	

	def set_name(self, name):
		self.name = name


	def set_rsd(self, rsd):
		self.rsd = rsd


	def calculate_avg_normarea(self, start_ind, end_ind):
		temp_normarea_total = 0
		avg_normarea = 0

		for i in range(start_ind, end_ind + 1):
			temp_normarea_total += self.data[i]

		avg_normarea = temp_normarea_total / (end_ind - start_ind + 1)
		self.avg_normarea = avg_normarea


class Metabolomics():
	def __init__(self, xlsx_filename):
		self.headers = Headers(xlsx_filename)
		self.metabolites = []
		self.repl_sample_names = []
		self.row_size = 0
		self.col_size = 0
		self.num_metabolites = 0
		self.autoset(xlsx_filename)
		self.autoset_repl_sample_names(xlsx_filename)


	def get_headers(self):
		return self.headers

	
	def get_row_size(self):
		return self.row_size

	
	def get_col_size(self):
		return self.col_size

	
	def get_metabolites(self):
		return self.metabolites


	def get_names(self):
		names = []
		for metabolite in self.metabolites:
			names.append(metabolite.get_name())
		return names


	def get_data_at_ind(self, ind):
		data = []
		for metabolite in self.metabolites:
			data.append(metabolite.get_data_at_ind(ind))
		return data


	def get_sample_name(self, ind):
		try:
			return self.repl_sample_names[ind]
		except Exception as e:
			print("get_sample_name: ", e, "\tindex: ", ind)
			return ""
			
	def get_num_metabolites(self):
		return self.num_metabolites


	def autoset(self, xlsx_filename):
		wb = openpyxl.load_workbook(xlsx_filename)
		ws = wb[WORKBOOK_SHEET_NAME]
		self.num_metabolites = ws.max_row - 1

		row_size = ws.max_row
		col_size = ws.max_column
		
		for row in range(2, ws.max_row + 1):
			metabolite_data = []
			for col in range(1, ws.max_column + 1):
				cell = ws.cell(row = row, column = col)
				metabolite_data.append(cell.value)

			self.metabolites.append(Metabolite(metabolite_data, self.headers.get_ind_name(), self.headers.get_ind_rsd(), self.headers.get_ind_normarea_start(), self.headers.get_ind_normarea_end()))


	def autoset_repl_sample_names(self, xlsx_filename):
		wb = openpyxl.load_workbook(xlsx_filename)
		ws = wb[WORKBOOK_SHEET_NAME]
		self.num_metabolites = ws.max_row - 1
		sample_name = ""
		temp_name = ""
		
		count = 0
		sample_num = 1
		for col in range(self.headers.get_ind_repl_start(), self.headers.get_ind_repl_end() + 1):
			cell = ws.cell(row = 1, column = col)
			for char in cell.value:
				if (count < 3):
					sample_name += char
					if (ord(char) < 65 or ord(char) > 90):
						count = 0
						sample_name = ""
						continue
					count += 1

			if (temp_name == sample_name):
				sample_num += 1
			else:
				sample_num = 1
				temp_name = sample_name

			sample_name += (" " + (str(sample_num).zfill(3)))
			self.repl_sample_names.append(sample_name)
			count = 0
			sample_name = ""

						

	def stitch_with(self, other):
		counter = 0
		for other_metab in other.get_metabolites():
			is_unique = True
			for self_metab in self.get_metabolites():
				if (other_metab == self_metab):
					is_unique = False
					if (other_metab > self_metab):
						self_metab = other_metab

			if (is_unique == True):
				self.metabolites.append(other_metab)
		
			counter += 1

		self.sort_list()


	def sort_list(self):
		sorted_names = []
		temp_metabolites = []

		for metabolite in self.get_metabolites():
			sorted_names.append(metabolite.get_name())
		sorted_names.sort()

		for name in sorted_names:
			for metabolite in self.get_metabolites():
				if (metabolite.get_name() == name):
					temp_metabolites.append(metabolite)
					break

		self.metabolites[:] = []
		self.metabolites = temp_metabolites

	
	def print_fancy(self):
		counter = 0
		for metabolite in self.metabolites:
			color = "\033[39;104m"
			if (counter % 2 == 0):
				color = "\033[39m"
			print(color, metabolite.get_name().ljust(35, ' '), str(metabolite.get_avg_normarea()).ljust(20, ' '), "\033[0m")
			counter += 1



def create_reformatted_xlsx_file(metabolomics, save_as):
	wb = Workbook()
	ws = wb.active
	ws.title = save_as
	header_names = ["Group", "Name"]

	for name in metabolomics.get_names():
		header_names.append(name)
	ws.append(header_names)

	counter = 0
	repl_ind_start = metabolomics.get_headers().get_ind_repl_start()
	repl_ind_end = metabolomics.get_headers().get_ind_repl_end() + 1
	for i in range(repl_ind_start - 1, repl_ind_end - 1):
		data = []
		data.append(metabolomics.get_sample_name(counter))
		data.append("PLACEHOLDER")
		for repl_data in metabolomics.get_data_at_ind(i):
			data.append(repl_data)
		ws.append(data)
		counter += 1

	dim_holder = DimensionHolder(worksheet=ws)
	for col in range(ws.min_column + 2, ws.max_column + 1):
		dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)		#MAGIC NUMBER
	ws.column_dimensions = dim_holder

	wb.save(save_as + ".xlsx")
	

def create_stitched_xlsx_file(pos, neg, stitched, save_as):
	def_cell_color = PatternFill(start_color="e26b0a", end_color="e26b0a", fill_type = "solid")	
	pos_cell_color = PatternFill(start_color="538dd5", end_color="538dd5", fill_type = "solid")	
	neg_cell_color = PatternFill(start_color="92d050", end_color="92d050", fill_type = "solid")	
	bold_font = Font(bold=True)
	data = []
	header = ["Positive Mode", "Negative Mode", "Both Modes", "Combined for MetaboAnalyst"]
	pos_counter = 0
	neg_counter = 0

	metabo_color_list = []

	for metabolite in stitched.get_metabolites():
		cell_color = def_cell_color
		is_pos = False
		is_neg = False
		line = ['','','','']

		pos_name = pos.get_metabolites()[pos_counter].get_name()
		neg_name = neg.get_metabolites()[neg_counter].get_name()
		pos_normarea = pos.get_metabolites()[pos_counter].get_avg_normarea()
		neg_normarea = neg.get_metabolites()[neg_counter].get_avg_normarea()

		if pos_name == metabolite.get_name():
			if pos_normarea == metabolite.get_avg_normarea():
				cell_color = pos_cell_color
			line[0] = metabolite.get_name()
			is_pos = True
			pos_counter += 1

		if neg_name == metabolite.get_name():
			if neg_normarea == metabolite.get_avg_normarea() and not pos_normarea == metabolite.get_avg_normarea():
				cell_color = neg_cell_color
			line[1] = metabolite.get_name()
			is_neg = True
			neg_counter += 1

		if is_pos and is_neg:
			line[2] = metabolite.get_name()
		line[3] = metabolite.get_name()
		data.append(line)
		metabo_color_list.append(cell_color)
	
	wb = Workbook()
	ws = wb.active
	ws.title = save_as
	ws.append(header)

	for col in ws["A:D"]:
		for cell in col:
			cell.font = bold_font

	for line in data:
		ws.append(line)

	counter = 0
	outline = Side(style = "thin", color = "000000")
	for row in ws.iter_rows(min_row=2, min_col = 4, max_col=4, max_row=ws.max_row):
		for cell in row:
			cell.fill = metabo_color_list[counter]
			cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
		counter += 1


	dim_holder = DimensionHolder(worksheet=ws)
	for col in range(ws.min_column, ws.max_column + 1):
		dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=30)
	dim_holder['F'] = ColumnDimension(ws, min=ws.max_column + 2, max=ws.max_column + 2, width=15)
	ws.column_dimensions = dim_holder

	cell = ws["F2"]
	cell.value = "positive"
	cell.fill = pos_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "right", vertical = "bottom")

	cell = ws["F3"]
	cell.value = "negative"
	cell.fill = neg_cell_color
	cell.border = Border(top=outline, left=outline, right=outline, bottom=outline)
	cell.alignment = Alignment(horizontal = "right", vertical = "bottom")

	wb.save(save_as + ".xlsx")



if __name__ == "__main__":
	metab_pos = Metabolomics("norm_data_pos.xlsx")
	metab_neg = Metabolomics("norm_data_neg.xlsx")
	metab_stitched = Metabolomics("norm_data_pos.xlsx")
	metab_stitched.stitch_with(Metabolomics("norm_data_neg.xlsx"))
	
	metab_stitched.print_fancy()
	create_reformatted_xlsx_file(metab_pos, "metab_pos_reformatted")
	create_reformatted_xlsx_file(metab_neg, "metab_neg_reformatted")
	create_stitched_xlsx_file(metab_pos, metab_neg, metab_stitched, "metab_stitched_table")
